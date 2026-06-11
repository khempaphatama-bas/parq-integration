from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path


OUT = Path("outputs/stakeholder_dashboard/PARQ_Stakeholder_Dashboard.xlsx")


COLORS = {
    "navy": "1F4E78",
    "blue": "5B9BD5",
    "light_blue": "D9EAF7",
    "green": "70AD47",
    "light_green": "E2F0D9",
    "amber": "FFC000",
    "light_amber": "FFF2CC",
    "red": "C00000",
    "light_red": "FCE4D6",
    "gray": "D9E1F2",
    "dark_gray": "404040",
    "white": "FFFFFF",
}


def style_sheet(ws, freeze="A2"):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )


def title(ws, text, subtitle=None, cols=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(1, 1, text)
    cell.font = Font(size=18, bold=True, color=COLORS["white"])
    cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
        scell = ws.cell(2, 1, subtitle)
        scell.font = Font(italic=True, color=COLORS["dark_gray"])
        scell.fill = PatternFill("solid", fgColor=COLORS["gray"])


def add_table(ws, name, start_row, headers, rows):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        cell.font = Font(bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r, row in enumerate(rows, start_row + 1):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    end_row = start_row + len(rows)
    end_col = len(headers)
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    return end_row


def widths(ws, mapping):
    for col, width in mapping.items():
        ws.column_dimensions[col].width = width


def status_fill(ws, status_col):
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, status_col)
        text = str(cell.value or "").lower()
        if "blocked" in text or "high" in text or "unavailable" in text:
            cell.fill = PatternFill("solid", fgColor=COLORS["light_red"])
        elif "open" in text or "pending" in text or "question" in text or "draft" in text:
            cell.fill = PatternFill("solid", fgColor=COLORS["light_amber"])
        elif "done" in text or "complete" in text or "current" in text or "recorded" in text:
            cell.fill = PatternFill("solid", fgColor=COLORS["light_green"])


deliverables = [
    ["PARQ-SOT-001", "PARQ Phase 1 User Flow Index", "Source of Truth", "TBD", "Current", "01_Source_of_Truth/PARQ_User_Flow/The_PARQ_Phase_1_User_Flow_Index.xlsx", "03_Architecture, QA, Product", "Primary feature/user-flow source."],
    ["PARQ-SOT-002", "PARQ Clarification Decision Log", "Source / Decision Log", "Libra", "BMS decision recorded / open questions", "01_Source_of_Truth/Clarifications/PARQ_Clarification_Decision_Log.md", "All agents", "BMS Option B recorded; proposal and owner gaps remain open."],
    ["PARQ-ARCH-001", "PARQ User Flow Integration Architecture", "Architecture", "TBD", "Current", "03_Architecture/PARQ_User_Flow_Integration_Architecture.md", "Simon, Quinn, Molly, PARQ", "Maps flows, dependencies, failure cases, and sequence candidates."],
    ["PARQ-ARCH-004", "PARQ Data/API/Boundary/Vendor Matrix", "Architecture", "TBD", "Current", "03_Architecture/PARQ_Data_API_Context_Boundary_Vendor_Matrix.md", "Simon, Quinn, PARQ", "Data ownership, API inventory, context, and vendor dependencies."],
    ["PARQ-ARCH-005", "PARQ Technical Dependency Control Pack", "Architecture / Control", "Simon", "Draft / Open Questions", "03_Architecture/PARQ_Technical_Dependency_Control_Pack.md", "Quinn, PARQ, Libra", "Vendor/API ownership, PII, error baseline, readiness inputs."],
    ["PARQ-ARCH-006", "PARQ Visual Architecture and Flow Pack", "Architecture / Visual", "Simon", "Draft / Open Questions", "03_Architecture/PARQ_Visual_Architecture_and_Flow_Pack.md", "Business, vendor, management, Quinn", "Visual context, boundaries, dependencies, sequences, risks."],
    ["PARQ-ARCH-007", "PARQ BMS Identity Flow Impact Assessment", "Architecture / Decision Impact", "Simon", "Draft / Option B Non-Blocking Decision Recorded", "03_Architecture/PARQ_BMS_Identity_Flow_Impact_Assessment.md", "PARQ, Molly, Quinn, Libra", "BMS login behavior decision recorded; technical controls remain open."],
    ["PARQ-UX-001", "PARQ UX Stakeholder User Flow Pack", "Discovery / UX", "Molly", "Draft / Open Questions", "02_Discovery/PARQ_UX_Stakeholder_User_Flow_Pack.md", "UX/UI, business, stakeholders", "Journey-first stakeholder flow artifact."],
    ["PARQ-REVIEW-001", "PARQ Integration Architecture Review", "Discovery / Review", "TBD", "Reference", "02_Discovery/PARQ_Integration_Architecture_Review.md", "Simon, Libra", "Historical-only vs active reference still unconfirmed."],
    ["PARQ-SYS-002", "BMS Member API Reference", "Source Reference", "TBD", "Current", "01_Source_of_Truth/API_and_System_References/00_2025_Document/api-members-by-account-id.md", "Simon, Quinn", "BMS GET /members/by-account-id source reference."],
    ["PARQ-SYS-003", "Add Identity Flow with BMS CheckMember", "Source Reference", "TBD", "Current", "01_Source_of_Truth/API_and_System_References/00_2025_Document/add_identity_flow.md", "Simon, Quinn, Molly", "Confirms current BMS usage in identity flow."],
    ["PARQ-MISSING-001", "Proposal: The PARQ integration.pdf", "Missing Source", "TBD", "Unavailable", "Not found in repository", "Libra, Simon, PARQ", "Must be uploaded/linked or formally accepted as unavailable."],
]

open_questions = [
    ["OQ-CLAR-001", "Where is [Proposal] The PARQ integration.pdf, or who can provide the approved source file/link?", "TBD", "Open", "MASTER_INDEX.md; PARQ_Clarification_Decision_Log.md", "Libra, PARQ, Simon", "Source traceability"],
    ["OQ-CLAR-002", "Is there an approved clarification note or decision record behind older architecture references?", "TBD", "Open", "PARQ_Clarification_Decision_Log.md", "All agents", "Source traceability"],
    ["OQ-CLAR-004", "Is PARQ_Integration_Architecture_Review.md historical-only or still an active reference?", "TBD", "Open", "MASTER_INDEX.md", "Libra, PARQ, Simon", "Document governance"],
    ["OQ-BMS-001", "What is the exact technical endpoint and payload contract for login-time BMS member check?", "IAM/BMS owners TBD", "Open", "PARQ_Clarification_Decision_Log.md; PARQ_BMS_Identity_Flow_Impact_Assessment.md", "Simon, Quinn", "BMS Option B"],
    ["OQ-BMS-002", "What timeout, retry, circuit-breaker, audit, and monitoring behavior applies to non-blocking BMS login refresh?", "IAM/BMS owners TBD", "Open", "Technical Dependency Control Pack", "Simon, Quinn, PARQ", "BMS Option B"],
    ["OQ-BMS-003", "If BMS returns member-bound-to-another-account during login, what support message and correction path apply?", "PARQ/IAM/Support TBD", "Open", "Clarification Decision Log", "Molly, Quinn, Simon", "Support flow"],
    ["OQ-BMS-004", "What cache/source detects previous Workplace permission while BMS is unavailable, and what freshness limit applies?", "IAM/FS owners TBD", "Open", "Clarification Decision Log; Technical Dependency Control Pack", "Simon, Quinn", "Fallback behavior"],
    ["OQ-OWN-001", "Who are the named owners for each indexed folder and document?", "Project leads", "Open", "MASTER_INDEX.md", "Libra, PARQ", "Ownership"],
    ["OQ-API-001", "Who owns SSO, IAM, FS, BZB, Argento, CMS, Kafka/Event Bus, BMS, Notification, Elevator, Turnstile APIs?", "PARQ / system owners", "Open", "Technical Dependency Control Pack", "Simon, Quinn, PARQ", "Dependency ownership"],
    ["OQ-PII-001", "What are field-level PII, consent, retention, and deletion rules per system/API?", "Security/privacy owner TBD", "Open", "Technical Dependency Control Pack", "Simon, Quinn, Management", "Compliance"],
    ["OQ-ENV-001", "What SIT/UAT/PVT endpoints, credentials, data sets, hardware paths, and readiness dates are available?", "PARQ / Quinn / vendors", "Open", "Technical Dependency Control Pack", "Quinn, PARQ, Vendor team", "Environment readiness"],
    ["OQ-PAY-001", "What idempotency, reconciliation schedule, refund flow, and support path apply to Argento/FS/OBK parking payment?", "OBK Backend / Argento / FS TBD", "Open", "Technical Dependency Control Pack", "Simon, Quinn, Vendor team", "Payments"],
]

risks = [
    ["R-001", "Named owners for external APIs are not confirmed.", "High", "Dependency follow-up and defect triage may stall.", "Confirm named owner for each system/API.", "PARQ / project leads", "Open"],
    ["R-002", "Proposal PDF is missing from repository.", "Medium", "Source traceability gap for proposal-derived assumptions.", "Upload/link approved proposal or accept existing review as sufficient reference.", "PARQ / Libra", "Open"],
    ["R-003", "Clarification source is incomplete as standalone artifact.", "Medium", "Architecture may rely on clarification content that is not independently traceable.", "Add approved clarification source or decision entries.", "PARQ / Libra / Simon", "Open"],
    ["R-004", "PII field-level mapping is incomplete.", "High", "PDPA, consent, deletion, and vendor sharing risks remain uncontrolled.", "Create approved PII/consent field matrix per API/system.", "Security/privacy owner TBD", "Open"],
    ["R-005", "BZB conflict and wrong-merge correction process is not operationally defined.", "High", "Account merge errors may require ad hoc backend fixes.", "Define correction runbook, approval, and audit trail.", "IAM/BZB/support owners TBD", "Open"],
    ["R-006", "FS is a multi-service dependency without per-service SLA/error contract.", "High", "Parking, visitor, elevator, turnstile, workplace, and traffic may fail inconsistently.", "Split FS API ownership and define SLA, timeout, retry, error catalog per service.", "FS owner TBD", "Open"],
    ["R-007", "QR single-use/replay behavior is not confirmed.", "Medium", "Access control may be weaker or inconsistent for visitors.", "Decide reusable-window versus single-use policy.", "FS / security / access owner TBD", "Open"],
    ["R-008", "FS outage behavior for turnstile/elevator is undefined.", "High", "Physical access may fail during network/system outage.", "Define online validation outage policy and site fallback.", "FS / site operations TBD", "Open"],
    ["R-009", "Argento/FS/OBK payment reconciliation details are incomplete.", "High", "Paid-but-unsynced parking tickets and refund disputes may occur.", "Define idempotency, reconciliation, refund, and support escalation.", "OBK Backend / Argento / FS TBD", "Open"],
    ["R-010", "CMS Phase 1 seed account may expose broader OBK data.", "High", "Data isolation and operational access risk.", "Decide compensating control or risk acceptance.", "CMS / security / PARQ TBD", "Open"],
    ["R-011", "Notification segment ownership and cleanup verification are not defined.", "Medium", "Wrong audience or stale data after deletion.", "Assign segment owner and cleanup validation.", "Notification / IAM TBD", "Open"],
    ["R-012", "Kafka/Event Bus retry and dead-letter behavior is unknown.", "Medium", "Permanent-delete cleanup may silently fail.", "Confirm topic schema, consumers, retry, DLQ, replay.", "Event platform TBD", "Open"],
    ["R-013", "SIT/UAT/PVT readiness lacks endpoint, data, and hardware dates.", "High", "Quinn cannot finalize planning or dependency gating.", "Provide environment readiness tracker with owners and dates.", "PARQ / Quinn / vendors", "Open"],
    ["R-014", "Phase 1.5 OCR/rate config dependencies may cause rework if deferred too far.", "Medium", "Parking e-stamp and fee calculation may need retrofit.", "Reserve Phase 1.5 data model and ownership decisions early.", "PARQ / Simon / FS / CMS / TCCT TBD", "Open"],
    ["R-015", "BMS non-blocking login refresh has unresolved controls.", "High", "Users may enter app but see stale/missing Workplace persona or require unclear support path.", "Define endpoint/payload, timeout, circuit-breaker, cache/source, support, PII, audit.", "IAM / BMS / PARQ / support TBD", "Open"],
]

dependencies = [
    ["SSO", "Authentication and identity orchestration", "High", "TBD", "POST /user/exists; POST /oauth/token; conflict behavior", "Users cannot login/register if unavailable.", "Owner, token/error contract, environment readiness"],
    ["IAM", "Local member profile and persona entitlement", "High", "TBD", "GET /identity/validate; POST /auth/login; Account Merge; BMS orchestration", "Persona and account lifecycle become unreliable.", "Profile overwrite policy, deletion orchestration, BMS mutation rules"],
    ["BMS", "Member lookup, non-blocking login refresh, cleanup consumer", "Medium/High", "TBD", "GET /members/by-account-id; checkMember; deletion cleanup", "Persona freshness may degrade; cleanup/compliance risk.", "Endpoint/payload, timeout, previous-permission source, audit, PII"],
    ["FS / Iviva", "Workplace, parking, visitor, elevator, turnstile, traffic authority", "High", "TBD", "Entitlements, floors, parking, visitor, QR validation, hardware", "Access, parking, visitor, elevator, turnstile can fail.", "Split service owners, SLAs, error catalog, hardware readiness"],
    ["BZB / Buzzee Bee", "Retail identity and privilege source", "High for merge", "TBD", "Email/phone lookup, BZB member ID, linkage sync", "Retail persona merge may fail or mislink.", "Duplicate handling, wrong-merge correction, deferrable behavior"],
    ["Argento", "Payment gateway and refund execution", "High for parking payment", "TBD", "PromptPay initiation, callback, refund", "Parking payment unavailable or unsynced.", "Callback security, idempotency, reconciliation, refund process"],
    ["CMS", "Admin visibility and operational tooling", "Medium / Phase 1.5 High", "TBD", "User view, visitor visibility, notification audience, car park config", "Operational support and data isolation gaps.", "Seed admin scope, RBAC/org filtering, compensating controls"],
    ["Kafka / Event Bus", "Async cleanup/event distribution", "Compliance-critical", "TBD", "ob-iam.account.permanent-deleted", "Hard-delete cleanup may fail silently.", "Schema, producer/consumer SLA, retry, DLQ, replay"],
    ["Notification Infrastructure", "PARQ targeting and device/inbox cleanup", "Medium", "TBD", "Segments, device tokens, inbox records", "Wrong audience or stale tokens after delete.", "Segment owner, token cleanup, validation"],
    ["Elevator System", "Physical access hardware", "High", "TBD", "Elevator command/status through FS", "Authorized access may fail at site.", "Hardware test environment, floor mapping, timeout/fallback"],
    ["Turnstile System", "Physical access hardware", "High", "TBD", "QR scan and online FS validation", "Access may be denied/allowed incorrectly.", "Outage policy, latency threshold, audit retention"],
    ["TCCT OCR", "Phase 1.5 ticket recognition", "Deferred / High later", "TBD", "OCR endpoint, ticket image/token, confidence", "Automated e-stamp cannot launch.", "API owner, sample data, image retention, confidence thresholds"],
]

overview_rows = [
    ["Dashboard purpose", "Stakeholder-friendly view of PARQ Integration status, deliverables, questions, risks, and dependencies."],
    ["Audience", "Product Team, Business Team, Vendor Team, Management"],
    ["Source basis", "Repository files only. No new requirements were created."],
    ["Recommended tabs", "Project Overview; Deliverables Tracker; Open Questions Tracker; Risk Register; Dependency Tracker"],
    ["Latest repository date observed", "2026-06-11"],
    ["Key current decision", "BMS Option B: login-time BMS check is non-blocking; app entry continues if BMS is unavailable."],
    ["Primary open management item", "Named owners and environment readiness remain unconfirmed for several vendor/system dependencies."],
]

wb = Workbook()
ws = wb.active
ws.title = "Project Overview"
title(ws, "PARQ Integration Stakeholder Dashboard", "Readable without opening markdown files", 8)
for idx, (label, value) in enumerate(overview_rows, 4):
    ws.cell(idx, 1, label).font = Font(bold=True, color=COLORS["navy"])
    ws.cell(idx, 2, value)
add_table(ws, "Overview_KPIs", 13, ["Metric", "Value", "Notes"], [
    ["Tracked deliverables", len(deliverables), "From MASTER_INDEX.md"],
    ["Open questions", len(open_questions), "From clarification log and technical control pack"],
    ["Open risks", len(risks), "From technical dependency control pack"],
    ["Tracked dependencies", len(dependencies), "Vendor/system dependencies"],
    ["Missing source artifacts", 1, "Proposal PDF unavailable"],
])
add_table(ws, "Overview_Tabs", 21, ["Recommended Tab", "Purpose", "Main Users"], [
    ["Project Overview", "Executive summary and dashboard key points", "Management, all teams"],
    ["Deliverables Tracker", "Artifact inventory, owner, status, source path, downstream consumer", "Product, PM, Management"],
    ["Open Questions Tracker", "Decision and information gaps requiring owner response", "Business, Product, Vendor"],
    ["Risk Register", "High-level project risks and mitigation needs", "Management, PM, Vendor"],
    ["Dependency Tracker", "System/vendor dependency visibility and readiness gaps", "Vendor, QA, Architecture"],
])
widths(ws, {"A": 28, "B": 85, "C": 38, "D": 22, "E": 22, "F": 30, "G": 30, "H": 30})
style_sheet(ws, "A4")

chart = BarChart()
chart.title = "Dashboard Counts"
chart.y_axis.title = "Count"
chart.x_axis.title = "Metric"
data = Reference(ws, min_col=2, min_row=13, max_row=17)
cats = Reference(ws, min_col=1, min_row=14, max_row=17)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 7
chart.width = 14
ws.add_chart(chart, "D13")

sheets = [
    ("Deliverables Tracker", "Deliverables_Tracker", ["ID", "Deliverable", "Category", "Owner", "Status", "Source / Input File", "Downstream Consumer", "Notes"], deliverables, {"A": 16, "B": 36, "C": 24, "D": 18, "E": 28, "F": 60, "G": 36, "H": 60}, 5),
    ("Open Questions Tracker", "Open_Questions", ["ID", "Open Question", "Owner", "Status", "Source / Input File", "Downstream Consumer", "Theme"], open_questions, {"A": 16, "B": 72, "C": 28, "D": 16, "E": 52, "F": 34, "G": 24}, 4),
    ("Risk Register", "Risk_Register", ["ID", "Risk", "Severity", "Impact", "Mitigation / Required Decision", "Owner", "Status"], risks, {"A": 14, "B": 46, "C": 16, "D": 56, "E": 60, "F": 34, "G": 16}, 7),
    ("Dependency Tracker", "Dependency_Tracker", ["System / Vendor", "Role", "Criticality", "Owner", "Interfaces / Data", "Impact if Unready", "Open Items"], dependencies, {"A": 24, "B": 46, "C": 20, "D": 18, "E": 52, "F": 50, "G": 58}, 4),
]

for sheet_name, table_name, headers, rows, width_map, status_col in sheets:
    ws = wb.create_sheet(sheet_name)
    title(ws, sheet_name, "Stakeholder-readable tracker generated from repository source files", len(headers))
    add_table(ws, table_name, 4, headers, rows)
    widths(ws, width_map)
    style_sheet(ws, "A5")
    status_fill(ws, status_col)
    if sheet_name == "Risk Register":
        for row in range(5, ws.max_row + 1):
            if str(ws.cell(row, 3).value).lower() == "high":
                ws.cell(row, 3).fill = PatternFill("solid", fgColor=COLORS["light_red"])
            elif str(ws.cell(row, 3).value).lower() == "medium":
                ws.cell(row, 3).fill = PatternFill("solid", fgColor=COLORS["light_amber"])

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if cell.row in (1, 2):
                cell.border = Border()
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)

# Verification pass.
check = load_workbook(OUT, data_only=False)
required = ["Project Overview", "Deliverables Tracker", "Open Questions Tracker", "Risk Register", "Dependency Tracker"]
missing = [s for s in required if s not in check.sheetnames]
if missing:
    raise RuntimeError(f"Missing sheets: {missing}")
for s in required:
    if check[s].max_row < 5:
        raise RuntimeError(f"Sheet appears underpopulated: {s}")
print(OUT)
